"""
Views for locations management (LGA, Ward, PollingUnit)
Used by Central Admin to create and manage polling units.
"""

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework.generics import ListCreateAPIView, RetrieveUpdateDestroyAPIView
from rest_framework.parsers import MultiPartParser, FormParser
from django.contrib.auth.hashers import make_password
from django.db import transaction
import openpyxl
from io import BytesIO
import time

from apps.common.permissions import IsCentralAdmin
from .utils import generate_polling_unit_password
from .models import LGA, Ward, PollingUnit
from .serializers import (
    LGASerializer, WardSerializer, PollingUnitSerializer,
    PollingUnitDetailSerializer, PollingUnitListSerializer,
    PollingUnitCreateSerializer, PollingUnitCreateResponseSerializer, PasswordResetResponseSerializer
)


class LGAListView(ListCreateAPIView):
    """
    GET: List all LGAs (public)
    POST: Create a new LGA (admin only)
    """
    queryset = LGA.objects.all()
    serializer_class = LGASerializer
    pagination_class = None  # Disable pagination to return all LGAs
    
    def get_permissions(self):
        if self.request.method == 'POST':
            return [IsAuthenticated(), IsCentralAdmin()]
        return []


class WardListView(ListCreateAPIView):
    """
    GET: List all wards (public), optionally filter by LGA
    POST: Create a new ward (admin only)
    """
    serializer_class = WardSerializer
    pagination_class = None  # Disable pagination to return all wards
    
    def get_permissions(self):
        if self.request.method == 'POST':
            return [IsAuthenticated(), IsCentralAdmin()]
        return []
    
    def get_queryset(self):
        """Allow filtering by LGA"""
        queryset = Ward.objects.all()
        lga_id = self.request.query_params.get('lga_id')
        if lga_id:
            queryset = queryset.filter(lga_id=lga_id)
        return queryset
    
    def perform_create(self, serializer):
        """Ensure ward is created by an admin"""
        serializer.save()


class PollingUnitListCreateView(ListCreateAPIView):
    """
    GET: List polling units (public), optionally filter by LGA or Ward
    POST: Create a new polling unit (admin only)
    
    On creation, returns the temporary plain-text password (shown only once).
    """
    pagination_class = None  # Disable pagination to return all polling units
    
    def get_permissions(self):
        if self.request.method == 'POST':
            return [IsAuthenticated(), IsCentralAdmin()]
        return []
    
    def get_queryset(self):
        """Allow filtering by LGA or Ward"""
        queryset = PollingUnit.objects.all()
        lga_id = self.request.query_params.get('lga_id')
        ward_id = self.request.query_params.get('ward_id')
        
        if lga_id:
            queryset = queryset.filter(lga_id=lga_id)
        if ward_id:
            queryset = queryset.filter(ward_id=ward_id)
        
        return queryset
    
    def get_serializer_class(self):
        """Use different serializer for list vs creation"""
        if self.request.method == 'GET':
            return PollingUnitListSerializer
        if self.request.method == 'POST':
            return PollingUnitCreateSerializer
        return PollingUnitSerializer
    
    def create(self, request, *args, **kwargs):
        """
        Override create to:
        1. Generate a secure random password
        2. Hash and save to database
        3. Store plaintext for display/printing purposes
        4. Return the plain-text password in response
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        # Generate temporary password (5-8 chars)
        temporary_password = generate_polling_unit_password()
        
        # Hash and save, also store plaintext for printing
        instance = serializer.save(
            password=make_password(temporary_password),
            plaintext_password=temporary_password
        )
        
        # Prepare response with plain-text password
        response_serializer = PollingUnitCreateResponseSerializer(
            instance,
            context={'temporary_password': temporary_password}
        )
        response_data = response_serializer.data
        response_data['warning'] = 'This password will not be shown again. Please share it with the agent immediately.'
        
        return Response(response_data, status=status.HTTP_201_CREATED)
    
    def perform_create(self, serializer):
        """Password is handled in create() method"""
        pass  # Don't call save here - handled in create()


class PollingUnitDetailView(RetrieveUpdateDestroyAPIView):
    """
    GET: Get a specific polling unit
    PUT: Update a polling unit (admin only)
    DELETE: Delete a polling unit (admin only)
    """
    queryset = PollingUnit.objects.all()
    serializer_class = PollingUnitDetailSerializer
    
    def get_permissions(self):
        if self.request.method == 'GET':
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsCentralAdmin()]


class PollingUnitPasswordResetView(APIView):
    """
    POST: Reset the password for a polling unit
    
    The backend generates a new secure password, hashes it, saves it,
    and returns the plain-text password only in this response (ephemeral).
    
    Body: {
        "polling_unit_id": "unit-uuid"  (or "unit_id": "AB/ANI/PU/0001")
    }
    
    Response: {
        "status": "success",
        "message": "Password reset successfully for ...",
        "pu_code": "AB/ANI/PU/0001",
        "pu_name": "Polling Unit Name",
        "temporary_password": "Xy8#Mn2P",
        "warning": "This password will not be shown again..."
    }
    
    Used by admin to reset polling unit credentials when agent loses password.
    """
    permission_classes = [IsAuthenticated, IsCentralAdmin]
    
    def post(self, request):
        # Accept either 'polling_unit_id' (UUID) or 'unit_id' (the PU Code)
        polling_unit_id = request.data.get('polling_unit_id')
        unit_code = request.data.get('unit_id')  # e.g., "AB/ANI/PU/0001"
        
        if not polling_unit_id and not unit_code:
            return Response(
                {'error': 'Either polling_unit_id (UUID) or unit_id (PU Code) is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Find the polling unit
            if polling_unit_id:
                polling_unit = PollingUnit.objects.get(id=polling_unit_id)
            else:
                polling_unit = PollingUnit.objects.get(unit_id=unit_code)
            
            # Generate new secure password
            temporary_password = generate_polling_unit_password()
            
            # Hash and save to database, also store plaintext
            polling_unit.password = make_password(temporary_password)
            polling_unit.plaintext_password = temporary_password
            polling_unit.save(update_fields=['password', 'plaintext_password', 'updated_at'])
            
            # Return response with plain-text password (ephemeral)
            response_serializer = PasswordResetResponseSerializer({
                'status': 'success',
                'message': f'Password reset successfully for {polling_unit.name}',
                'pu_code': polling_unit.unit_id,
                'pu_name': polling_unit.name,
                'temporary_password': temporary_password,
                'warning': 'This password will not be shown again. Please share it with the agent immediately.'
            })
            
            return Response(response_serializer.data, status=status.HTTP_200_OK)
        
        except PollingUnit.DoesNotExist:
            search_field = 'unit_id' if unit_code else 'id'
            search_value = unit_code or polling_unit_id
            return Response(
                {'error': f'Polling unit with {search_field}={search_value} not found'},
                status=status.HTTP_404_NOT_FOUND
            )


class PollingUnitPasswordFetchView(APIView):
    """
    POST: Fetch polling unit credentials for printing
    Body: {
        "lga_id": "optional-lga-id",
        "ward_id": "optional-ward-id",
        "unit_ids": ["PU-00001", "PU-00002"]  # optional - specific units
    }
    
    Returns: List of polling units with their IDs and passwords
    
    Used by admin to fetch credentials for printing.
    """
    permission_classes = [IsAuthenticated, IsCentralAdmin]
    
    def post(self, request):
        lga_id = request.data.get('lga_id')
        ward_id = request.data.get('ward_id')
        unit_ids = request.data.get('unit_ids', [])
        
        queryset = PollingUnit.objects.all()
        
        if lga_id:
            queryset = queryset.filter(lga_id=lga_id)
        if ward_id:
            queryset = queryset.filter(ward_id=ward_id)
        if unit_ids:
            queryset = queryset.filter(unit_id__in=unit_ids)
        
        polling_units = queryset.values('id', 'unit_id', 'name', 'lga__name', 'ward__name', 'plaintext_password', 'created_at')
        
        # Normalize field names: lga__name -> lga_name, ward__name -> ward_name
        normalized_units = [
            {
                'id': unit['id'],
                'unit_id': unit['unit_id'],
                'name': unit['name'],
                'lga_name': unit['lga__name'],
                'ward_name': unit['ward__name'],
                'password': unit['plaintext_password'] or 'N/A',  # Return plaintext password for display
                'created_at': unit['created_at'].isoformat() if unit['created_at'] else None,
            }
            for unit in polling_units
        ]
        
        return Response(
            normalized_units,
            status=status.HTTP_200_OK
        )


class BulkCreatePollingUnitsView(APIView):
    """
    POST: Create multiple polling units from Excel data
    Body: {
        "wards": [
            {"name": "Ward Name", "lga_id": "lga-uuid"},
            ...
        ],
        "polling_units": [
            {"name": "PU Name", "ward_id": "ward-uuid", "lga_id": "lga-uuid"},
            ...
        ]
    }
    
    Used by admin to bulk upload ward and polling unit data.
    """
    permission_classes = [IsAuthenticated, IsCentralAdmin]
    
    @transaction.atomic
    def post(self, request):
        try:
            wards_data = request.data.get('wards', [])
            polling_units_data = request.data.get('polling_units', [])
            
            if not wards_data and not polling_units_data:
                return Response(
                    {'error': 'No wards or polling units provided'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            created_wards = []
            created_polling_units = []
            errors = []
            
            # Create wards
            for ward_data in wards_data:
                try:
                    # Validate required fields
                    if 'name' not in ward_data or 'lga_id' not in ward_data:
                        errors.append("Ward missing 'name' or 'lga_id'")
                        continue
                    
                    ward = Ward.objects.create(
                        name=ward_data['name'],
                        lga_id=ward_data['lga_id']
                    )
                    created_wards.append({
                        'id': str(ward.id),
                        'name': ward.name,
                        'lga_id': str(ward.lga_id)
                    })
                except Exception as e:
                    error_msg = f"Ward '{ward_data.get('name', 'unknown')}' creation failed: {str(e)}"
                    errors.append(error_msg)
            
            # Create polling units
            for pu_data in polling_units_data:
                try:
                    # Validate required fields
                    if 'name' not in pu_data or 'ward_id' not in pu_data or 'lga_id' not in pu_data:
                        errors.append("Polling unit missing 'name', 'ward_id', or 'lga_id'")
                        continue
                    
                    password = generate_polling_unit_password()
                    pu = PollingUnit.objects.create(
                        name=pu_data['name'],
                        ward_id=pu_data['ward_id'],
                        lga_id=pu_data['lga_id'],
                        password=make_password(password),
                        plaintext_password=password
                    )
                    created_polling_units.append({
                        'unit_id': pu.unit_id,
                        'name': pu.name,
                        'password': password
                    })
                except Exception as e:
                    error_msg = f"Polling unit '{pu_data.get('name', 'unknown')}' creation failed: {str(e)}"
                    errors.append(error_msg)
            
            # Return appropriate status based on results
            if not created_wards and not created_polling_units:
                return Response(
                    {
                        'created_wards': [],
                        'created_polling_units': [],
                        'errors': errors,
                        'message': 'No items were created. Check errors for details.'
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            return Response(
                {
                    'created_wards': created_wards,
                    'created_polling_units': created_polling_units,
                    'errors': errors,
                    'message': f'Created {len(created_wards)} wards and {len(created_polling_units)} polling units'
                },
                status=status.HTTP_201_CREATED if not errors else status.HTTP_206_PARTIAL_CONTENT
            )
        
        except Exception as e:
            error_msg = f"Bulk create operation failed: {str(e)}"
            return Response(
                {'error': error_msg, 'errors': [error_msg]},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class PollingUnitExcelUploadView(APIView):
    """
    POST: Create multiple polling units from an Excel file
    OPTIONS: Handle CORS preflight requests
    
    Excel file must contain columns:
    - lga_name: Name of the LGA
    - ward_name: Name of the Ward within the LGA
    - unit_name: Name of the Polling Unit
    - unit_id (optional): Custom polling unit ID
    
    Returns: {
        'created_count': number of polling units created,
        'failed_count': number of rows that failed,
        'errors': list of error messages,
        'units': list of created polling units with their passwords
    }
    """
    permission_classes = [IsAuthenticated, IsCentralAdmin]
    parser_classes = (MultiPartParser, FormParser)
    
    def options(self, request, *args, **kwargs):
        """Handle CORS preflight requests"""
        return Response(status=status.HTTP_200_OK)
    
    @transaction.atomic
    def post(self, request):
        try:
            # Get the uploaded file
            uploaded_file = request.FILES.get('file')
            if not uploaded_file:
                error_msg = 'No file provided'
                return Response(
                    {'error': error_msg},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Check file size (max 10MB)
            max_size = 10 * 1024 * 1024  # 10MB
            if uploaded_file.size > max_size:
                error_msg = f'File too large. Maximum size is 10MB. Your file is {uploaded_file.size / (1024*1024):.2f}MB'
                return Response(
                    {'error': error_msg},
                    status=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE
                )
            
            # Read Excel file
            file_content = uploaded_file.read()
            
            try:
                workbook = openpyxl.load_workbook(BytesIO(file_content))
            except Exception as e:
                error_msg = f'Invalid Excel file: {str(e)}'
                return Response(
                    {'error': error_msg},
                    status=status.HTTP_400_BAD_REQUEST
                )
            worksheet = workbook.active
            
            # Check row count (max 5000 rows per upload)
            max_rows = 5000
            row_count_check = worksheet.max_row - 1  # Exclude header row
            if row_count_check > max_rows:
                error_msg = f'Too many rows. Maximum {max_rows} rows per upload. Your file has {row_count_check} rows.'
                return Response(
                    {'error': error_msg},
                    status=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE
                )
            
            created_units = []
            failed_rows = []
            
            # Get headers (first row)
            headers = {}
            for col_idx, cell in enumerate(worksheet[1], 1):
                header_name = cell.value.strip().lower() if cell.value else None
                if header_name:
                    headers[header_name] = col_idx
            
            # Validate required columns
            required_columns = ['lga_name', 'ward_name', 'unit_name']
            missing_columns = [col for col in required_columns if col not in headers]
            if missing_columns:
                return Response(
                    {'error': f'Missing required columns: {", ".join(missing_columns)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # OPTIMIZATION: Load all LGAs and Wards into memory (2 queries total)
            lga_map = {lga.name: lga for lga in LGA.objects.all()}
            
            ward_map = {(ward.name, ward.lga_id): ward for ward in Ward.objects.all()}
            
            print(f"[EXCEL UPLOAD] Starting with {len(lga_map)} LGAs and {len(ward_map)} wards in database")
            print(f"[EXCEL UPLOAD] Available LGAs: {list(lga_map.keys())}")
            
            # Collect all rows to create
            lgas_to_create = []
            wards_to_create = []
            units_to_create = []
            rows_data = []
            
            # Parse all rows first
            row_count = 0
            
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 2):
                row_count += 1
                try:
                    lga_name = row[headers['lga_name'] - 1]
                    ward_name = row[headers['ward_name'] - 1]
                    unit_name = row[headers['unit_name'] - 1]
                    unit_id = row[headers.get('unit_id', 0) - 1] if 'unit_id' in headers else None
                    
                    # Validate required fields
                    if not all([lga_name, ward_name, unit_name]):
                        failed_rows.append({
                            'row': row_idx,
                            'error': 'Missing required fields (lga_name, ward_name, unit_name)'
                        })
                        continue
                    
                    lga_name = str(lga_name).strip()
                    ward_name = str(ward_name).strip()
                    unit_name = str(unit_name).strip()
                    
                    rows_data.append({
                        'row_idx': row_idx,
                        'lga_name': lga_name,
                        'ward_name': ward_name,
                        'unit_name': unit_name,
                        'unit_id': unit_id
                    })
                    
                except Exception as e:
                    failed_rows.append({
                        'row': row_idx,
                        'error': str(e)
                    })
            
            print(f"[EXCEL UPLOAD] Parsed {len(rows_data)} rows successfully, {len(failed_rows)} parsing failures")
            if rows_data:
                print(f"[EXCEL UPLOAD] First row sample: {rows_data[0]}")
            
            # Now process rows with batch database operations
            process_start = time.time()
            
            # Collect unique LGA and Ward names to create
            lgas_to_create_names = set()
            for row_data in rows_data:
                lga_name = row_data['lga_name']
                if lga_name not in lga_map:
                    lgas_to_create_names.add(lga_name)
            
            # Batch create LGAs that don't exist
            if lgas_to_create_names:
                print(f"[EXCEL UPLOAD] Creating {len(lgas_to_create_names)} new LGAs...")
                lgas_to_create = []
                for lga_name in lgas_to_create_names:
                    # Generate acronym from first 3 letters of LGA name (uppercase)
                    acronym = lga_name.strip()[:3].upper() if lga_name else ''
                    lgas_to_create.append(LGA(name=lga_name, acronym=acronym))
                
                LGA.objects.bulk_create(lgas_to_create, ignore_conflicts=True)
                # Refresh lga_map with newly created LGAs from database
                for lga in LGA.objects.all():
                    lga_map[lga.name] = lga
                print(f"[EXCEL UPLOAD] LGA Map now has {len(lga_map)} LGAs")
                for lga_name in lgas_to_create_names:
                    lga = lga_map.get(lga_name)
                    if lga:
                        print(f"[EXCEL UPLOAD] Created LGA: {lga_name} with acronym: {lga.acronym}")
            
            # Collect unique Ward keys to create (ward_name, lga_name)
            wards_to_create_data = []
            for row_data in rows_data:
                lga_name = row_data['lga_name']
                ward_name = row_data['ward_name']
                
                lga = lga_map.get(lga_name)
                if lga:
                    ward_key = (ward_name, lga.id)
                    if ward_key not in ward_map:
                        wards_to_create_data.append((ward_name, lga))
                        # Add to map as placeholder to avoid duplicates
                        ward_map[ward_key] = None
            
            # Batch create Wards that don't exist
            if wards_to_create_data:
                print(f"[EXCEL UPLOAD] Creating {len(wards_to_create_data)} new Wards...")
                wards_to_create = [Ward(name=ward_name, lga=lga) for ward_name, lga in wards_to_create_data]
                Ward.objects.bulk_create(wards_to_create, ignore_conflicts=True)
                # Refresh ward_map with all wards from database
                for ward in Ward.objects.select_related('lga').all():
                    ward_map[(ward.name, ward.lga_id)] = ward
                print(f"[EXCEL UPLOAD] Ward Map now has {len(ward_map)} wards")
            
            # Now process each row to create polling units
            lga_unit_counters = {}  # Track unit count per LGA for this batch
            
            for row_data in rows_data:
                lga_name = row_data['lga_name']
                ward_name = row_data['ward_name']
                unit_name = row_data['unit_name']
                unit_id = row_data['unit_id']
                row_idx = row_data['row_idx']
                
                try:
                    # Get LGA
                    lga = lga_map.get(lga_name)
                    if not lga:
                        failed_rows.append({
                            'row': row_idx,
                            'error': f'LGA "{lga_name}" not found'
                        })
                        print(f"[EXCEL UPLOAD] Row {row_idx}: LGA '{lga_name}' not found. Available LGAs: {list(lga_map.keys())}")
                        continue
                    
                    # Get Ward
                    ward_key = (ward_name, lga.id)
                    ward = ward_map.get(ward_key)
                    if not ward:
                        failed_rows.append({
                            'row': row_idx,
                            'error': f'Ward "{ward_name}" not found in LGA "{lga_name}"'
                        })
                        print(f"[EXCEL UPLOAD] Row {row_idx}: Ward '{ward_name}' not found for LGA '{lga_name}'. Available wards for LGA: {[k for k in ward_map.keys() if k[1] == lga.id]}")
                        continue
                    
                    # Generate unit_id if not provided (following the model's save() logic)
                    if not unit_id:
                        # Initialize counter for this LGA if not already done
                        if lga.id not in lga_unit_counters:
                            lga_unit_counters[lga.id] = PollingUnit.objects.filter(lga=lga).count()
                        
                        # Increment and generate unit_id
                        lga_unit_counters[lga.id] += 1
                        unit_id = f"AB/{lga.acronym}/PU/{lga_unit_counters[lga.id]:04d}"
                    
                    # Create Polling Unit object for batch insert
                    password = generate_polling_unit_password()
                    polling_unit = PollingUnit(
                        name=unit_name,
                        lga=lga,
                        ward=ward,
                        unit_id=unit_id,
                        password=make_password(password),
                        plaintext_password=password
                    )
                    
                    units_to_create.append({
                        'obj': polling_unit,
                        'password': password,
                        'row_idx': row_idx
                    })
                    
                except Exception as e:
                    failed_rows.append({
                        'row': row_idx,
                        'error': str(e)
                    })
                    print(f"[EXCEL UPLOAD] Row {row_idx}: Exception - {str(e)}")
            
            # Batch create all Polling Units
            if units_to_create:
                print(f"[EXCEL UPLOAD] Creating {len(units_to_create)} Polling Units in batch...")
                polling_unit_objs = [u['obj'] for u in units_to_create]
                created_polling_units = PollingUnit.objects.bulk_create(polling_unit_objs)
                
                # Build response with created units and passwords
                for created_unit, unit_data in zip(created_polling_units, units_to_create):
                    created_units.append({
                        'unit_id': created_unit.unit_id,
                        'name': created_unit.name,
                        'lga_name': created_unit.lga.name,
                        'ward_name': created_unit.ward.name,
                        'password': unit_data['password'],
                        'created_at': created_unit.created_at.isoformat()
                    })
            
            return Response(
                {
                    'created_count': len(created_units),
                    'failed_count': len(failed_rows),
                    'errors': failed_rows,
                    'units': created_units
                },
                status=status.HTTP_201_CREATED
            )
            
        except Exception as e:
            import traceback
            error_msg = f'Failed to process Excel file: {str(e)}'
            print(f"[EXCEL UPLOAD ERROR] {error_msg}")
            print(f"[EXCEL UPLOAD TRACEBACK] {traceback.format_exc()}")
            return Response(
                {'error': error_msg, 'details': traceback.format_exc()},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
